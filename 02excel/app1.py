from openpyxl import Workbook
# from openpyxl import load_workbook


# 加载一个已有的文件
# wb = load_workbook('.xlsx')
# 获取所有sheet
# sheets = wb.sheetnames
# 选择到一个sheet
# ws = wb['sheetname']

# 创建一个新文件
wb = Workbook()
ws = wb.create_sheet(title='sheet1', index=0)

ws.append([1,2,3,4])
wb.save('02excel/test.xlsx')

