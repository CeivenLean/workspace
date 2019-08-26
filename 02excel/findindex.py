import os
from openpyxl import load_workbook


wb = load_workbook('/Users/xxxl/VSProjects/workspace/02excel/test.xlsx')
sheet = wb.active
def get_index(path):
    for cells in sheet.iter_cols(): 
        # print(cells)
        for cell2 in cells: 
            if cell2.value is not None: 
                # print(cell2.value)
                # print(type(cell2.value))
                info2 = str(cell2.value).find('3')
                if info2 == 0: 
                    print(cell2.row,cell2.column)

if __name__ == "__main__":
    print(os.getcwd())  # /Users/xxxl/VSProjects/workspace

    path1=os.path.abspath('.')   #表示当前所处的文件夹的绝对路径
    path2=os.path.abspath('..')
    print(path1)  # /Users/xxxl/VSProjects/workspace
    print(path2)  # /Users/xxxl/VSProjects
    d = os.path.dirname(__file__)
    print(d)  # /Users/xxxl/VSProjects/workspace/02excel
    d2 = os.path.abspath(__file__)
    print(d2)  # /Users/xxxl/VSProjects/workspace/02excel/findindex.py

                
